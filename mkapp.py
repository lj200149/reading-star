# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import openpyxl,os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
import sys,types
from pypinyin import pinyin, lazy_pinyin, Style
reload(sys)
sys.setdefaultencoding('utf-8')
fw=open("makeapp.bash","w")
fy=open("en.yml","w")
fy.write("en:\n")
f_index=open("index.html.erb","w")
dirs=os.listdir('.')
f_index.write("<h1>")
views=[]
for f in dirs:
   if f.split(".")[-1]=="xlsx":
       appname=f.split(".")[0]
       f_index.write(appname+"!</h1>\n")
       app_py=lazy_pinyin(appname)
       py=""
       for p in app_py:
           py=py + p
       print appname, py
       fy.write("\n  "+py+": '"+appname+"'")
       fw.write("rails new "+py+"  -f \n")
       fw.write("cd "+py+"\n")
       fw.write("rails generate controller Welcome index\n")
       app_py=py
       wb = load_workbook(filename = f)
       for sheet in wb:
           sheetname=sheet.title
           sheet_py=lazy_pinyin(sheetname,errors='replace')
           py=""
           for p in sheet_py:
               py=py + p
           print " ",sheetname, py
           sheetname_py=py
           views.append(py+"s")
           fy.write("\n  "+py+": '"+sheetname+"'")
           fw.write("rails generate scaffold "+py)
           f_index.write("<%= link_to '")
           f_index.write(sheetname)
           f_index.write("', controller: '")
           f_index.write(py)
           f_index.write("s' %>\n")
           ws=wb[sheetname]
           row_range = ws[1:1]
           for cells in row_range:
               comment = cells.comment
               if comment:print comment
               if type(cells.value) is types.UnicodeType:
                   field=cells.value.replace("\n","")
                   field_py=lazy_pinyin(field,errors='replace')
                   py=""
                   for p in field_py:
                       py=py + p #.capitalize()
                   print "   ",field, py
                   fw.write(" "+py+":string")
                   fy.write("\n  "+py+": '"+field+"'")
           fw.write("  --skip-stylesheets\n")

fw.write("rails db:migrate\n")
fw.write("sed -i '1a\  get \"welcome/index\"' config/routes.rb\n")
fw.write("sed -i '9a\  root \"welcome#index\"' config/routes.rb\n")
#fw.write("cd ..\n")
fw.write("cp  ../index.html.erb ")
#fw.write(app_py)
fw.write("app/views/welcome/index.html.erb\n")

fw.write("sed -i \"7a\gem \'twitter-bootstrap-rails\'\" Gemfile\n")
fw.write("sed -i \"8a\gem \'therubyracer\'\" Gemfile\n")
fw.write("sed -i \"9a\gem \'less-rails\'\" Gemfile\n")

fw.write("bundle install\n")

fw.write("rails g bootstrap:install\n")

for l in views:
    fw.write("rails g bootstrap:themed "+l+" -f\n")

fw.write("cp -f ../en.yml ")
#fw.write(app_py)
fw.write("config/locales/en.yml\n")
dirs= "app/views/"

for l in views:
    f=dirs+l+"/_form.html.erb"
    fw.write("sed -i 's/f.label/f.label t/' "+f+"\n")

#fw.write("mv app/assets/stylesheets/application.css app/assets/stylesheets/application.scss\n")
#fw.write("sed -i '15a\// \"bootstrap-sprockets\" must be imported before \"bootstrap\" and \"bootstrap/variables\"' app/assets/stylesheets/application.scss\n")
#fw.write("sed -i '16a\@import \"bootstrap-sprockets\";' app/assets/stylesheets/application.scss\n")
#fw.write("sed -i '17a\@import \"bootstrap\";' app/assets/stylesheets/application.scss\n")




#fw.write("bundle install\n")
fw.write("rails server\n")
fw.write("firefox localhost:3000\n")
fw.close()
fy.close()
f_index.close()
